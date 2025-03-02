from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from colorful_training_template.excel_generator.label_creator import LabelCreator

FONT_SIZE_10_BOLD = Font(size=10, bold=True)
FONT_SIZE_8 = Font(size=8)


class DataPopulator:
    def __init__(self, sheet):
        self.sheet = sheet

    def populate_workout_data(
        self, start_row, start_col, week_data, num_sets, set_width, num_exercises
    ):
        # Define mapping from weekday names to indices (0 = Monday, ..., 6 = Sunday)
        WEEKDAY_TO_INDEX = {
            "Monday": 0,
            "Tuesday": 1,
            "Wednesday": 2,
            "Thursday": 3,
            "Friday": 4,
            "Saturday": 5,
            "Sunday": 6,
        }
        # Build a mapping from weekday index to the corresponding day data.
        weekday_to_data = {}
        for day in week_data:
            weekday = day.get("weekday")
            if weekday in WEEKDAY_TO_INDEX:
                index = WEEKDAY_TO_INDEX[weekday]
                weekday_to_data[index] = day

        # Iterate over all 7 days so that each day block is in its proper position.
        for day_idx in range(7):
            # Calculate the starting row for this day block.
            day_block_start = start_row + day_idx * (num_exercises + 1)
            day_data = weekday_to_data.get(day_idx)
            if day_data is not None:
                for exercise_idx, exercise_data in enumerate(
                    day_data.get("exercises", [])
                ):
                    if exercise_idx < num_exercises:
                        self._populate_exercise_data(
                            day_block_start + exercise_idx,
                            start_col + 3,
                            exercise_data,
                            num_sets,
                            set_width,
                        )
            else:
                # Optionally, clear or leave blank the block for missing days.
                pass

    def _populate_exercise_data(
        self, start_row, start_col, exercise_data, num_sets, set_width
    ):
        self._populate_exercise_name(start_row, start_col, exercise_data["name"])
        for set_idx, set_data in enumerate(exercise_data["sets"]):
            self._populate_set_data(
                start_row, start_col + 3 + set_idx * (set_width + 1), set_data
            )

    def _populate_exercise_name(self, start_row, start_col, name):
        cell = self.sheet.cell(row=start_row, column=start_col)
        cell.value = name
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = FONT_SIZE_10_BOLD
        LabelCreator(self.sheet).add_border_to_label(
            start_row, start_col, start_row, start_col
        )

    def _populate_set_data(self, start_row, start_col, set_data):
        self._populate_cell(start_row, start_col, set_data.get("reps", ""), "right")
        self._populate_cell(
            start_row, start_col + 1, set_data.get("weight", ""), "right"
        )
        self._populate_percentage_cell(
            start_row, start_col + 2, set_data.get("percentage_1rm", "")
        )
        self._populate_cell(
            start_row, start_col + 3, set_data.get("notes", ""), "left", FONT_SIZE_8
        )

    def _populate_cell(self, row, col, value, align, font=None):
        cell = self.sheet.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if font:
            cell.font = font

        # If the value is a string, check its length against the current column width.
        if isinstance(value, str):
            col_letter = get_column_letter(col)
            # Get the current column width (or default to Excel's default if not set).
            current_width = self.sheet.column_dimensions[col_letter].width
            if current_width is None:
                current_width = 8.43  # Default width in Excel
            # If the string's length exceeds the current width, adjust the column.
            if len(value) > current_width:
                auto_adjust_column_width(self.sheet, col)

    def _populate_percentage_cell(self, row, col, value):
        cell = self.sheet.cell(row=row, column=col)
        if value in (None, ""):
            cell.value = ""
        else:
            cell.value = value / 100
            cell.number_format = "0%" if value == int(value) else "0.0%"
        cell.alignment = Alignment(horizontal="right", vertical="center")


def auto_adjust_column_width(sheet, col_index, scaling_factor=0.7, padding=0):
    """
    Automatically adjust the width of a column based on the maximum length
    of its cell values. The width is calculated as:

        width = (max length of cell values) * scaling_factor + padding

    You can tweak `scaling_factor` and `padding` to reduce or increase the extra space.
    """
    column_letter = get_column_letter(col_index)
    max_length = 0
    for cell in sheet[column_letter]:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    sheet.column_dimensions[column_letter].width = max_length * scaling_factor + padding
