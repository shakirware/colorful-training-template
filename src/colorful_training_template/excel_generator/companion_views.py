from __future__ import annotations

from datetime import datetime, timedelta
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

WEEKDAY_TO_DAY_NUM = {
    "Monday": 1,
    "Tuesday": 2,
    "Wednesday": 3,
    "Thursday": 4,
    "Friday": 5,
    "Saturday": 6,
    "Sunday": 7,
}

DAY_LABELS = [f"Day {i}" for i in range(1, 8)]

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FILL = PatternFill("solid", fgColor="65BCCC")
SECTION_FILL = PatternFill("solid", fgColor="8CCDD8")
HEADER_FILL = PatternFill("solid", fgColor="79C4D2")
CONTROL_FILL = PatternFill("solid", fgColor="EEF8FA")
DATA_FILL = PatternFill("solid", fgColor="F7FCFD")
MUTED_FILL = PatternFill("solid", fgColor="E8F5F7")

TITLE_FONT = Font(size=14, bold=True)
SECTION_FONT = Font(size=12, bold=True)
HEADER_FONT = Font(bold=True)
BODY_FONT = Font(size=11)
BODY_BOLD_FONT = Font(size=11, bold=True)
SMALL_FONT = Font(size=10, italic=True)
SMALL_BOLD_FONT = Font(size=10, bold=True)


class CompanionViewError(Exception):
    """Raised when workbook companion views cannot be created."""


def add_companion_views_to_workbook(
    workbook,
    calculated_program: list[dict[str, Any]],
    start_date: datetime,
) -> None:
    rows, week_labels, max_exercises = _flatten_program(calculated_program, start_date)

    for sheet_name in ("Week View", "Today", "_Program_Data"):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    week_view = workbook.create_sheet("Week View")
    today_view = workbook.create_sheet("Today")
    program_data = workbook.create_sheet("_Program_Data")

    _populate_program_data_sheet(program_data, rows)
    _build_week_view_sheet(week_view, max_exercises, week_labels)
    _build_today_sheet(today_view, max_exercises, week_labels)

    program_data.sheet_state = "hidden"

    _reorder_sheets(workbook)
    workbook.active = workbook.sheetnames.index("Week View")


def _flatten_program(
    calculated_program: list[dict[str, Any]],
    start_date: datetime,
) -> tuple[list[dict[str, Any]], list[str], int]:
    rows: list[dict[str, Any]] = []
    week_labels: list[str] = []
    max_exercises = 1

    for week_index, week_dict in enumerate(calculated_program, start=1):
        week_start = start_date + timedelta(weeks=week_index - 1)
        week_label = _week_label(week_index, week_start)
        week_labels.append(week_label)

        if not isinstance(week_dict, dict):
            raise CompanionViewError(
                f"Each week must be a mapping, got {type(week_dict).__name__}"
            )

        week_days = next(iter(week_dict.values()), [])
        for day in week_days:
            weekday = day.get("weekday")
            if weekday not in WEEKDAY_TO_DAY_NUM:
                continue

            day_num = WEEKDAY_TO_DAY_NUM[weekday]
            exercises = day.get("exercises", [])
            max_exercises = max(max_exercises, len(exercises) or 1)

            for exercise_index, exercise in enumerate(exercises, start=1):
                rows.append(
                    {
                        "week_num": week_index,
                        "week_label": week_label,
                        "week_start": week_start.strftime("%d-%m-%Y"),
                        "day_num": day_num,
                        "day_label": f"Day {day_num}",
                        "exercise_index": exercise_index,
                        "exercise": exercise.get("name", ""),
                        "prescription": _build_prescription_summary(
                            exercise.get("sets", [])
                        ),
                        "notes": _build_notes_summary(exercise.get("sets", [])),
                        "key": week_index * 1000 + day_num * 100 + exercise_index,
                    }
                )

    return rows, week_labels, max_exercises


def _build_prescription_summary(sets: list[dict[str, Any]]) -> str:
    groups: list[tuple[int, Any, Any, Any]] = []

    for set_data in sets:
        reps = set_data.get("reps")
        weight = set_data.get("weight")
        percentage = set_data.get("percentage_1rm")

        if groups and groups[-1][1:] == (reps, weight, percentage):
            count, _, _, _ = groups[-1]
            groups[-1] = (count + 1, reps, weight, percentage)
        else:
            groups.append((1, reps, weight, percentage))

    parts: list[str] = []
    for count, reps, weight, percentage in groups:
        if reps in (None, ""):
            base = ""
        else:
            base = f"{count}×{reps}"

        if weight not in (None, ""):
            base = f"{base} @ {weight}" if base else str(weight)

        if percentage not in (None, ""):
            pct_text = _format_percentage(percentage)
            base = f"{base} ({pct_text}%)" if base else f"{pct_text}%"

        if base:
            parts.append(base)

    return " • ".join(parts)


def _build_notes_summary(sets: list[dict[str, Any]]) -> str:
    note_groups: list[str] = []

    for set_data in sets:
        note = str(set_data.get("notes", "") or "").strip()
        if not note:
            continue
        if not note_groups or note_groups[-1] != note:
            note_groups.append(note)

    return " | ".join(note_groups)


def _format_percentage(value: Any) -> str:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)

    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.1f}".rstrip("0").rstrip(".")


def _week_label(week_num: int, week_start: datetime) -> str:
    return f"Week {week_num} - Week Commencing {week_start.strftime('%d-%m-%Y')}"


def _populate_program_data_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Week Num",
        "Week Label",
        "Week Start",
        "Day Num",
        "Day Label",
        "Exercise Index",
        "Exercise",
        "Prescription",
        "Notes",
        "Key",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append(
            [
                row["week_num"],
                row["week_label"],
                row["week_start"],
                row["day_num"],
                row["day_label"],
                row["exercise_index"],
                row["exercise"],
                row["prescription"],
                row["notes"],
                row["key"],
            ]
        )

    widths = {
        "A": 10,
        "B": 34,
        "C": 14,
        "D": 10,
        "E": 10,
        "F": 14,
        "G": 28,
        "H": 58,
        "I": 80,
        "J": 12,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _build_week_view_sheet(sheet, max_exercises: int, week_labels: list[str]) -> None:
    _configure_companion_sheet(sheet)

    week_count = len(week_labels)

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Week View"
    _style_cell(sheet["A1"], fill=TITLE_FILL, font=TITLE_FONT, align="center")

    sheet["A2"] = "Selected week #"
    _style_cell(sheet["A2"], fill=CONTROL_FILL, font=HEADER_FONT)

    sheet["B2"] = 1 if week_count > 0 else ""
    _style_cell(
        sheet["B2"],
        fill=CONTROL_FILL,
        font=BODY_BOLD_FONT,
        border=True,
        align="center",
        vertical="center",
    )

    if week_count > 0:
        sheet["C2"] = (
            f"=IFERROR(VLOOKUP(B2,$E$4:$F${week_count + 3},2,FALSE),"
            f'"Type 1-{week_count}")'
        )
    else:
        sheet["C2"] = "No weeks available"

    _style_cell(sheet["C2"], fill=CONTROL_FILL, font=BODY_FONT, border=True)

    sheet.merge_cells("A3:C3")
    if week_count > 0:
        sheet["A3"] = f"Type a week number in B2. Valid range: 1 to {week_count}."
    else:
        sheet["A3"] = "No weeks available."
    _style_cell(sheet["A3"], fill=MUTED_FILL, font=SMALL_FONT)

    _write_week_legend(sheet, week_labels, start_row=3, start_col=5)

    if week_count > 0:
        _add_integer_validation(
            sheet,
            "B2",
            minimum=1,
            maximum=week_count,
            prompt="Enter a week number",
            error=f"Please enter a whole number from 1 to {week_count}.",
        )

    block_height = max_exercises + 3
    start_row = 5

    for day_num in range(1, 8):
        day_row = start_row + (day_num - 1) * block_height
        sheet.merge_cells(
            start_row=day_row,
            start_column=1,
            end_row=day_row,
            end_column=3,
        )
        sheet.cell(day_row, 1).value = f"Day {day_num}"
        _style_cell(sheet.cell(day_row, 1), fill=SECTION_FILL, font=SECTION_FONT)

        headers_row = day_row + 1
        for col, title in enumerate(["Exercise", "Plan", "Notes"], start=1):
            cell = sheet.cell(headers_row, col, value=title)
            _style_cell(
                cell,
                fill=HEADER_FILL,
                font=HEADER_FONT,
                align="center",
                vertical="center",
            )

        first_data_row = day_row + 2
        for exercise_index in range(1, max_exercises + 1):
            row = first_data_row + exercise_index - 1
            key_expr = f"IFERROR($B$2*1000+{day_num}*100+{exercise_index},0)"
            data_range_key = "_Program_Data!$J$2:$J$1048576"
            data_range_ex = "_Program_Data!$G$2:$G$1048576"
            data_range_plan = "_Program_Data!$H$2:$H$1048576"
            data_range_notes = "_Program_Data!$I$2:$I$1048576"

            if exercise_index == 1:
                exercise_formula = (
                    f'=IF(COUNTIF({data_range_key},{key_expr})=0,"Rest / no session",'
                    f'IFERROR(INDEX({data_range_ex},MATCH({key_expr},{data_range_key},0)),""))'
                )
                plan_formula = (
                    f'=IF(A{row}="Rest / no session","—",'
                    f'IFERROR(INDEX({data_range_plan},MATCH({key_expr},{data_range_key},0)),""))'
                )
            else:
                exercise_formula = f'=IFERROR(INDEX({data_range_ex},MATCH({key_expr},{data_range_key},0)),"")'
                plan_formula = f'=IFERROR(INDEX({data_range_plan},MATCH({key_expr},{data_range_key},0)),"")'

            notes_formula = f'=IFERROR(INDEX({data_range_notes},MATCH({key_expr},{data_range_key},0)),"")'

            sheet.cell(row, 1, value=exercise_formula)
            sheet.cell(row, 2, value=plan_formula)
            sheet.cell(row, 3, value=notes_formula)
            _style_data_row(sheet, row, emphasize_first=(exercise_index == 1))

    sheet.freeze_panes = "A5"


def _build_today_sheet(sheet, max_exercises: int, week_labels: list[str]) -> None:
    _configure_companion_sheet(sheet)

    week_count = len(week_labels)

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Today / Single-Day View"
    _style_cell(sheet["A1"], fill=TITLE_FILL, font=TITLE_FONT, align="center")

    sheet["A2"] = "Selected week #"
    _style_cell(sheet["A2"], fill=CONTROL_FILL, font=HEADER_FONT)

    sheet["B2"] = 1 if week_count > 0 else ""
    _style_cell(
        sheet["B2"],
        fill=CONTROL_FILL,
        font=BODY_BOLD_FONT,
        border=True,
        align="center",
        vertical="center",
    )

    if week_count > 0:
        sheet["C2"] = (
            f"=IFERROR(VLOOKUP(B2,$E$4:$F${week_count + 3},2,FALSE),"
            f'"Type 1-{week_count}")'
        )
    else:
        sheet["C2"] = "No weeks available"
    _style_cell(sheet["C2"], fill=CONTROL_FILL, font=BODY_FONT, border=True)

    sheet["A3"] = "Selected day #"
    _style_cell(sheet["A3"], fill=CONTROL_FILL, font=HEADER_FONT)

    sheet["B3"] = 1
    _style_cell(
        sheet["B3"],
        fill=CONTROL_FILL,
        font=BODY_BOLD_FONT,
        border=True,
        align="center",
        vertical="center",
    )

    sheet["C3"] = '=IFERROR(IF(AND(B3>=1,B3<=7),"Day "&B3,"Type 1-7"),"Type 1-7")'
    _style_cell(sheet["C3"], fill=CONTROL_FILL, font=BODY_FONT, border=True)

    sheet.merge_cells("A4:C4")
    if week_count > 0:
        sheet["A4"] = (
            f"Type a week number in B2 (1 to {week_count}) and a day number in B3 (1 to 7)."
        )
    else:
        sheet["A4"] = "No weeks available."
    _style_cell(sheet["A4"], fill=MUTED_FILL, font=SMALL_FONT)

    _write_week_legend(sheet, week_labels, start_row=3, start_col=5)
    _write_day_legend(sheet, start_row=week_count + 6, start_col=5)

    if week_count > 0:
        _add_integer_validation(
            sheet,
            "B2",
            minimum=1,
            maximum=week_count,
            prompt="Enter a week number",
            error=f"Please enter a whole number from 1 to {week_count}.",
        )

    _add_integer_validation(
        sheet,
        "B3",
        minimum=1,
        maximum=7,
        prompt="Enter a day number",
        error="Please enter a whole number from 1 to 7.",
    )

    sheet.merge_cells("A5:C5")
    sheet["A5"] = '=IFERROR(C2&" — "&C3,"Selected session")'
    _style_cell(sheet["A5"], fill=SECTION_FILL, font=SECTION_FONT)

    for col, title in enumerate(["Exercise", "Plan", "Notes"], start=1):
        cell = sheet.cell(6, col, value=title)
        _style_cell(
            cell,
            fill=HEADER_FILL,
            font=HEADER_FONT,
            align="center",
            vertical="center",
        )

    for exercise_index in range(1, max_exercises + 1):
        row = 6 + exercise_index
        key_expr = f"IFERROR($B$2*1000+$B$3*100+{exercise_index},0)"
        data_range_key = "_Program_Data!$J$2:$J$1048576"
        data_range_ex = "_Program_Data!$G$2:$G$1048576"
        data_range_plan = "_Program_Data!$H$2:$H$1048576"
        data_range_notes = "_Program_Data!$I$2:$I$1048576"

        if exercise_index == 1:
            exercise_formula = (
                f'=IF(COUNTIF({data_range_key},{key_expr})=0,"Rest / no session",'
                f'IFERROR(INDEX({data_range_ex},MATCH({key_expr},{data_range_key},0)),""))'
            )
            plan_formula = (
                f'=IF(A{row}="Rest / no session","—",'
                f'IFERROR(INDEX({data_range_plan},MATCH({key_expr},{data_range_key},0)),""))'
            )
        else:
            exercise_formula = f'=IFERROR(INDEX({data_range_ex},MATCH({key_expr},{data_range_key},0)),"")'
            plan_formula = f'=IFERROR(INDEX({data_range_plan},MATCH({key_expr},{data_range_key},0)),"")'

        notes_formula = f'=IFERROR(INDEX({data_range_notes},MATCH({key_expr},{data_range_key},0)),"")'

        sheet.cell(row, 1, value=exercise_formula)
        sheet.cell(row, 2, value=plan_formula)
        sheet.cell(row, 3, value=notes_formula)
        _style_data_row(sheet, row, emphasize_first=(exercise_index == 1), height=42)

    sheet.freeze_panes = "A6"


def _write_week_legend(
    sheet, week_labels: list[str], start_row: int, start_col: int
) -> None:
    number_col = start_col
    label_col = start_col + 1

    header_left = sheet.cell(start_row, number_col, value="Week #")
    header_right = sheet.cell(start_row, label_col, value="Week label")
    _style_cell(header_left, fill=HEADER_FILL, font=HEADER_FONT, align="center")
    _style_cell(header_right, fill=HEADER_FILL, font=HEADER_FONT, align="center")

    for idx, label in enumerate(week_labels, start=1):
        row = start_row + idx
        left = sheet.cell(row, number_col, value=idx)
        right = sheet.cell(row, label_col, value=label)
        _style_cell(left, fill=DATA_FILL, font=BODY_FONT, align="center")
        _style_cell(right, fill=DATA_FILL, font=BODY_FONT)


def _write_day_legend(sheet, start_row: int, start_col: int) -> None:
    number_col = start_col
    label_col = start_col + 1

    header_left = sheet.cell(start_row, number_col, value="Day #")
    header_right = sheet.cell(start_row, label_col, value="Day label")
    _style_cell(header_left, fill=HEADER_FILL, font=HEADER_FONT, align="center")
    _style_cell(header_right, fill=HEADER_FILL, font=HEADER_FONT, align="center")

    for idx, label in enumerate(DAY_LABELS, start=1):
        row = start_row + idx
        left = sheet.cell(row, number_col, value=idx)
        right = sheet.cell(row, label_col, value=label)
        _style_cell(left, fill=DATA_FILL, font=BODY_FONT, align="center")
        _style_cell(right, fill=DATA_FILL, font=BODY_FONT)


def _add_integer_validation(
    sheet,
    cell_ref: str,
    *,
    minimum: int,
    maximum: int,
    prompt: str,
    error: str,
) -> None:
    validation = DataValidation(
        type="whole",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=False,
    )
    validation.prompt = prompt
    validation.error = error
    sheet.add_data_validation(validation)
    validation.add(sheet[cell_ref])


def _configure_companion_sheet(sheet) -> None:
    widths = {
        "A": 20,
        "B": 12,
        "C": 52,
        "D": 3,
        "E": 10,
        "F": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.column_dimensions["D"].hidden = True
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24


def _style_data_row(
    sheet,
    row: int,
    emphasize_first: bool = False,
    height: int = 34,
) -> None:
    for col in range(1, 4):
        cell = sheet.cell(row, col)
        cell.fill = DATA_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = BODY_BOLD_FONT if (emphasize_first and col == 1) else BODY_FONT

    sheet.row_dimensions[row].height = height


def _style_cell(
    cell,
    *,
    fill=None,
    font=None,
    align: str | None = "left",
    vertical: str | None = None,
    border: bool = True,
) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border:
        cell.border = THIN_BORDER
    if align or vertical:
        cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=True)


def _reorder_sheets(workbook) -> None:
    primary_title = workbook.worksheets[0].title if workbook.worksheets else None
    ordered_titles: list[str] = []

    if primary_title:
        ordered_titles.append(primary_title)

    for title in ("Week View", "Today", "Charts", "_Program_Data"):
        if title in workbook.sheetnames and title not in ordered_titles:
            ordered_titles.append(title)

    for title in workbook.sheetnames:
        if title not in ordered_titles:
            ordered_titles.append(title)

    workbook._sheets = [workbook[title] for title in ordered_titles]
